"""
工资 API 端点模块

提供工资试算与查看的 RESTful API 端点：
- 查看自己的工资试算
- 管理员查看全员工资汇总
- 管理员设置工资参数
- 管理员导出工资表（Excel）
"""

import uuid
from datetime import datetime, timezone
from io import BytesIO
from typing import Annotated, Any, Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from app.core.dependencies import (
    CurrentUser,
    SessionDep,
    require_scope,
)
from app.core.scopes import SalaryScope
from app.core.schemas import Message
from app.core.errors import BusinessException, ErrorCode
from app.core.models import User

from app.domains.salary import repository
from app.domains.salary.schemas import (
    SalaryParamsUpdate,
    EngineerSalaryDetail,
    EngineerSalarySummary,
    PMSalaryDetail,
    PMSalarySummary,
    SalarySummaryList,
    SalaryExportRequest,
)
from app.domains.salary.service import calculate_all_salaries_detail, calculate_user_salary
from app.domains.audit.repository import create_audit_log


router = APIRouter()


# ==================== 工程师/PM 端点：工资试算 ====================


@router.get(
    "/my",
    response_model=EngineerSalaryDetail | PMSalaryDetail,
    summary="查看我的工资试算",
    description="工程师或 PM 查看自己的工资试算结果"
)
async def read_my_salary(
    session: SessionDep,
    current_user: CurrentUser,
    _: Annotated[None, Depends(require_scope(SalaryScope.READ))] = None,
) -> Any:
    """
    获取当前用户的工资试算

    权限：工程师或 PM（需 salary:read 权限）
    require_scope 已确保权限，calculate_user_salary 内部校验角色。

    计算规则：
    - 工程师：S下 = (S0 - P差额) × K
    - PM：S总 = S底 + S考
    """
    return await calculate_user_salary(session=session, user=current_user)


# ==================== 管理员端点：工资汇总和设置 ====================


@router.get(
    "",
    response_model=SalarySummaryList,
    summary="查看工资汇总（管理员）",
    description="管理员查看所有工程师和 PM 的工资汇总"
)
async def read_salary_summary(
    session: SessionDep,
    current_user: CurrentUser,
    page: Annotated[int, Query(ge=1, description="页码，从1开始")] = 1,
    page_size: Annotated[int, Query(ge=1, le=100, description="每页数量，默认20")] = 20,
    month: Annotated[Optional[str], Query(pattern=r"^\d{4}-\d{2}$", description="月份 YYYY-MM，默认当前月")] = None,
    _: Annotated[None, Depends(require_scope(SalaryScope.ADMIN))] = None,
) -> Any:
    """
    获取工资汇总列表

    权限：管理员（需 salary:admin 权限）
    """
    offset = (page - 1) * page_size

    # 获取用户列表
    users, count = await repository.get_all_salaries(
        session=session,
        skip=offset,
        limit=page_size,
    )

    # 计算工资明细
    result_salaries = await calculate_all_salaries_detail(
        session=session, users=users, month=month,
    )

    return SalarySummaryList(
        data=result_salaries,
        count=count,
        page=page,
        page_size=page_size,
        total_pages=(count + page_size - 1) // page_size if count > 0 else 0,
    )


@router.put(
    "/users/{user_id}/params",
    response_model=Message,
    summary="设置工资参数（管理员）",
    description="管理员设置工程师或 PM 的工资参数"
)
async def update_salary_params(
    *,
    session: SessionDep,
    current_user: CurrentUser,
    user_id: uuid.UUID,
    params: SalaryParamsUpdate,
    _: Annotated[None, Depends(require_scope(SalaryScope.ADMIN))] = None,
) -> Any:
    """
    设置工资参数

    权限：管理员（需 salary:admin 权限）

    可设置字段：
    - 工程师：S0, H0, T_monthly_plan
    - PM：S_base, S_assess, R_base, R_assess
    """
    # 转换为字典，过滤 None 值
    if not params.model_dump(exclude_none=True):
        raise BusinessException(
            code=ErrorCode.SYSTEM_VALIDATION_ERROR,
            detail="No parameters provided"
        )

    # 更新参数（直接传入 typed DTO，按角色过滤字段）
    user = await repository.update_user_salary_params(
        session=session,
        user_id=user_id,
        params=params,
    )

    if not user:
        raise BusinessException(
            code=ErrorCode.USER_NOT_FOUND,
            detail=f"User with id {user_id} not found"
        )

    await create_audit_log(
        session=session, user_id=current_user.id, action="salary.update",
        target_type="user", target_id=str(user_id),
        details=f"Salary parameters updated", ip_address=None,
    )

    return Message(message=f"Salary parameters updated for user {user_id}")


@router.post(
    "/export",
    summary="导出工资表（管理员）",
    description="管理员导出全员工资汇总表为 CSV 文件流"
)
async def export_salaries(
    *,
    session: SessionDep,
    current_user: CurrentUser,
    request: SalaryExportRequest,
    _: Annotated[None, Depends(require_scope(SalaryScope.ADMIN))] = None,
) -> Any:
    """
    导出工资表为 CSV

    权限：管理员（需 salary:admin 权限）
    """
    # 获取所有员工（先获取总数，避免硬编码截断）
    users, count = await repository.get_all_salaries(
        session=session,
        skip=0,
        limit=0,
    )
    if count > 0:
        users, _ = await repository.get_all_salaries(
            session=session,
            skip=0,
            limit=count,
        )

    # 计算每个员工的工资并生成 Excel
    result_salaries = await calculate_all_salaries_detail(
        session=session, users=users, month=request.month,
    )

    # 按角色分组
    engineer_rows = []
    pm_rows = []
    for s in result_salaries:
        if isinstance(s, EngineerSalarySummary):
            engineer_rows.append([
                s.full_name or "",
                s.S0,
                round(s.H0, 2) if s.H0 else "",
                s.T_monthly_plan or "",
                s.T_effective or "",
                s.T_actual_monthly or "",
                s.T_reported_monthly or "",
                round(s.P_diff, 2) if s.P_diff else "",
                s.k_coefficient or "",
                s.current_starpoint or "",
                round(s.salary_final, 2),
            ])
        else:
            pm_rows.append([
                s.full_name or "",
                s.S_base,
                s.S_assess,
                s.R_base or "",
                s.R_assess or "",
                round(s.salary_total, 2),
            ])

    # 生成 Excel
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def write_sheet(ws, title, headers, data_rows):
        ws.title = title
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        for row_idx, row_data in enumerate(data_rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col_idx)].width = 16

    # 工程师表
    ws_eng = wb.active
    write_sheet(ws_eng, "工程师", ["姓名", "S0", "H0", "T月计划", "T有效", "本月实际工时", "本月报价工时", "P差额", "K系数", "当前星点", "最终工资"], engineer_rows)

    # PM 表
    ws_pm = wb.create_sheet()
    write_sheet(ws_pm, "市场产品PM", ["姓名", "S底", "S考", "R底", "R考", "总工资"], pm_rows)

    # 写入 BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    month = request.month or datetime.now(timezone.utc).strftime("%Y-%m")
    filename = f"salary_export_{month}.xlsx"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""},
    )