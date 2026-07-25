"""
工资 API 集成测试

测试工资模块的 RESTful API 端点：
- 查看自己的工资试算
- 管理员查看工资汇总
- 管理员设置工资参数
"""

import uuid
from datetime import datetime, timedelta, timezone

import pytest
from httpx import AsyncClient
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.models import User, UserRoleType, Task, TaskStatus, TaskType, StarPointRecord, JudgmentType
from app.core.security import get_password_hash, create_access_token
from app.core.scopes import SalaryScope, TaskScope, BidScope, ReportScope, StarPointScope, UserScope, ClientResourceScope, RuleScope


async def create_test_engineer(session: AsyncSession) -> User:
    """创建测试工程师用户（含工资参数）"""
    from app.core.models import Role, RoleScope, UserRole

    engineer = User(
        email=f"engineer_salary_api_{uuid.uuid4()}@test.com",
        hashed_password=get_password_hash("testpassword"),
        full_name="Test Engineer",
        role=UserRoleType.ENGINEER,
        is_active=True,
        S0=10000.0,
        H0=50.0,
        T_monthly_plan=160.0,
        current_starpoint=100,
    )
    session.add(engineer)
    await session.commit()

    # 创建角色并关联 scopes
    role = Role(name=f"engineer_role_{uuid.uuid4().hex[:8]}")
    session.add(role)
    await session.commit()

    user_role = UserRole(user_id=engineer.id, role_id=role.id)
    session.add(user_role)

    for scope in [
        TaskScope.READ,
        BidScope.CREATE, BidScope.UPDATE,
        ReportScope.CREATE, ReportScope.READ,
        StarPointScope.READ,
        SalaryScope.READ,
    ]:
        rs = RoleScope(scope=scope.value, role_id=role.id)
        session.add(rs)

    await session.commit()
    return engineer


async def create_test_admin(session: AsyncSession) -> User:
    """创建测试管理员用户"""
    from app.core.models import Role, RoleScope, UserRole

    admin = User(
        email=f"admin_salary_api_{uuid.uuid4()}@test.com",
        hashed_password=get_password_hash("testpassword"),
        full_name="Test Admin",
        role=UserRoleType.ADMIN,
        is_active=True,
    )
    session.add(admin)
    await session.commit()

    role = Role(name=f"admin_{uuid.uuid4().hex[:8]}")
    session.add(role)
    await session.commit()

    user_role = UserRole(user_id=admin.id, role_id=role.id)
    session.add(user_role)

    for scope in [
        TaskScope.READ, TaskScope.CREATE, TaskScope.UPDATE, TaskScope.ADMIN,
        BidScope.READ,
        ReportScope.READ, ReportScope.ADMIN,
        StarPointScope.READ, StarPointScope.ADMIN,
        SalaryScope.READ, SalaryScope.ADMIN,
        ClientResourceScope.READ,
        UserScope.READ, UserScope.CREATE, UserScope.UPDATE, UserScope.ADMIN,
        RuleScope.ADMIN,
    ]:
        rs = RoleScope(scope=scope.value, role_id=role.id)
        session.add(rs)

    await session.commit()
    return admin


class TestSalaryAPIs:

    @pytest.mark.asyncio
    async def test_read_my_salary_engineer(self, client: AsyncClient, db_session: AsyncSession):
        """测试工程师查看自己的工资试算"""
        engineer = await create_test_engineer(db_session)

        token = create_access_token(
            subject=str(engineer.id),
            expires_delta=timedelta(minutes=30),
        )

        response = await client.get(
            "/v1/salaries/my",
            headers={"Authorization": f"Bearer {token}"},
        )

        assert response.status_code == 200
        data = response.json()
        assert data["role"] == "engineer"
        assert data["S0"] == 10000.0
        # H0 现在是自动计算 = S0 / T_monthly_plan = 10000 / 160 = 62.5
        assert data["H0"] == 62.5
        assert data["current_starpoint"] == 100
        assert "salary_final" in data

    @pytest.mark.asyncio
    async def test_read_my_salary_unauthorized(self, client: AsyncClient, db_session: AsyncSession):
        """测试未认证用户无法查看工资"""
        response = await client.get("/v1/salaries/my")
        assert response.status_code == 401

    @pytest.mark.asyncio
    async def test_read_salary_summary_as_admin(self, client: AsyncClient, db_session: AsyncSession):
        """测试管理员查看工资汇总"""
        admin = await create_test_admin(db_session)
        await create_test_engineer(db_session)  # 确保至少有一个工程师

        token = create_access_token(
            subject=str(admin.id),
            expires_delta=timedelta(minutes=30),
        )

        response = await client.get(
            "/v1/salaries",
            headers={"Authorization": f"Bearer {token}"},
        )

        assert response.status_code == 200
        data = response.json()
        assert isinstance(data["data"], list)
        assert data["count"] >= 1

    @pytest.mark.asyncio
    async def test_engineer_cannot_view_salary_summary(self, client: AsyncClient, db_session: AsyncSession):
        """测试工程师无法查看工资汇总（无 salary:admin 权限）"""
        engineer = await create_test_engineer(db_session)

        token = create_access_token(
            subject=str(engineer.id),
            expires_delta=timedelta(minutes=30),
        )

        response = await client.get(
            "/v1/salaries",
            headers={"Authorization": f"Bearer {token}"},
        )

        assert response.status_code == 403

    @pytest.mark.asyncio
    async def test_admin_update_salary_params(self, client: AsyncClient, db_session: AsyncSession):
        """测试管理员设置工资参数"""
        engineer = await create_test_engineer(db_session)
        admin = await create_test_admin(db_session)

        token = create_access_token(
            subject=str(admin.id),
            expires_delta=timedelta(minutes=30),
        )

        response = await client.put(
            f"/v1/salaries/users/{engineer.id}/params",
            json={
                "S0": 15000.0,
                "H0": 60.0,
            },
            headers={"Authorization": f"Bearer {token}"},
        )

        assert response.status_code == 200
        data = response.json()
        assert "message" in data

        # 验证参数已更新
        updated = await db_session.get(User, engineer.id)
        assert updated.S0 == 15000.0
        assert updated.H0 == 60.0

    @pytest.mark.asyncio
    async def test_non_admin_cannot_update_salary_params(self, client: AsyncClient, db_session: AsyncSession):
        """测试非管理员无法设置工资参数"""
        engineer = await create_test_engineer(db_session)

        token = create_access_token(
            subject=str(engineer.id),
            expires_delta=timedelta(minutes=30),
        )

        response = await client.put(
            f"/v1/salaries/users/{engineer.id}/params",
            json={"S0": 15000.0},
            headers={"Authorization": f"Bearer {token}"},
        )

        assert response.status_code == 403

    @pytest.mark.asyncio
    async def test_update_salary_params_not_found(self, client: AsyncClient, db_session: AsyncSession):
        """测试更新不存在的用户"""
        admin = await create_test_admin(db_session)

        token = create_access_token(
            subject=str(admin.id),
            expires_delta=timedelta(minutes=30),
        )

        response = await client.put(
            f"/v1/salaries/users/{uuid.uuid4()}/params",
            json={"S0": 15000.0},
            headers={"Authorization": f"Bearer {token}"},
        )

        assert response.status_code == 404

    @pytest.mark.asyncio
    async def test_update_salary_params_empty(self, client: AsyncClient, db_session: AsyncSession):
        """测试提交空参数"""
        engineer = await create_test_engineer(db_session)
        admin = await create_test_admin(db_session)

        token = create_access_token(
            subject=str(admin.id),
            expires_delta=timedelta(minutes=30),
        )

        response = await client.put(
            f"/v1/salaries/users/{engineer.id}/params",
            json={},
            headers={"Authorization": f"Bearer {token}"},
        )

        assert response.status_code == 400

    @pytest.mark.asyncio
    async def test_salary_pagination(self, client: AsyncClient, db_session: AsyncSession):
        """测试工资汇总分页"""
        admin = await create_test_admin(db_session)

        token = create_access_token(
            subject=str(admin.id),
            expires_delta=timedelta(minutes=30),
        )

        response = await client.get(
            "/v1/salaries?page=1&page_size=2",
            headers={"Authorization": f"Bearer {token}"},
        )

        assert response.status_code == 200
        data = response.json()
        assert "data" in data
        assert data["page"] == 1
        assert data["page_size"] == 2

    @pytest.mark.asyncio
    async def test_export_salaries(self, client: AsyncClient, db_session: AsyncSession):
        """测试导出工资表"""
        admin = await create_test_admin(db_session)
        await create_test_engineer(db_session)  # 确保至少有一个工程师
        # 创建 PM 用户，验证 PM 工资也能导出
        pm_user = User(
            email=f"pm_salary_api_{uuid.uuid4()}@test.com",
            hashed_password=get_password_hash("testpassword"),
            full_name="Test PM",
            role=UserRoleType.PM,
            is_active=True,
            S_base=8000.0,
            S_assess=2000.0,
            R_base=1.0,
            R_assess=1.0,
        )
        db_session.add(pm_user)
        await db_session.commit()

        token = create_access_token(
            subject=str(admin.id),
            expires_delta=timedelta(minutes=30),
        )

        response = await client.post(
            "/v1/salaries/export",
            json={
                "month": "2026-07",
            },
            headers={"Authorization": f"Bearer {token}"},
        )

        assert response.status_code == 200
        assert response.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "Content-Disposition" in response.headers
        assert "salary_export" in response.headers["content-disposition"]
        assert ".xlsx" in response.headers["content-disposition"]

        # 验证 Excel 内容包含工程师和 PM 两个 sheet
        from io import BytesIO
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(response.content))
        assert "工程师" in wb.sheetnames
        assert "市场产品PM" in wb.sheetnames
        pm_sheet = wb["市场产品PM"]
        # 表头: 姓名, S底, S考, R底, R考, 总工资
        assert pm_sheet.cell(1, 1).value == "姓名"
        assert pm_sheet.cell(1, 2).value == "S底"
        # 第二行是数据行，验证 PM 用户出现在表格中
        pm_names = [pm_sheet.cell(r, 1).value for r in range(2, pm_sheet.max_row + 1)]
        assert "Test PM" in pm_names, f"PM 'Test PM' not found in export: {pm_names}"
